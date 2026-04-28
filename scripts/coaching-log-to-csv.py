"""Flatten /docs/Coaching log.xlsx into a single CSV matching the
ARTES /api/coaching/hours/import schema.

Each tab has slightly different columns, so this script applies a
per-sheet mapping. Output columns match the importer:

    date, hours, category, client_type, paid_status,
    client_name, client_organization, mentor_coach_name, notes

Usage:
    python scripts/coaching-log-to-csv.py docs/Coaching\\ log.xlsx \
        docs/coaching-log-import.csv
"""
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


# ─── Field normalizers ────────────────────────────────────────────────────────

def normalize_date(v):
    """Return ISO YYYY-MM-DD or '' if unparseable."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    # Try DD/MM/YYYY (the dominant string format in this spreadsheet).
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def normalize_hours(v):
    """Convert a duration cell to a decimal-hour number, or None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    if not s:
        return None
    # "45 mins" / "45 minutes" / "30 min"
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(min|mins|minute|minutes)\b", s)
    if m:
        return round(float(m.group(1)) / 60.0, 2)
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(h|hr|hrs|hour|hours)?$", s)
    if m:
        return float(m.group(1))
    return None


def clean_name(v):
    return str(v).strip() if v is not None else ""


# ─── Per-sheet handlers ───────────────────────────────────────────────────────
#
# Each handler is a generator that yields dicts matching the importer's
# expected column names. Handlers receive the raw row (a tuple) and the
# zero-based row index, and skip blanks themselves.

def rows_paid_private(row, idx):
    name, date, company, dur, _ = (list(row) + [None] * 5)[:5]
    if not name or not date or dur is None: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": clean_name(company),
        "notes": "",
    }


def rows_treasury_board(row, idx):
    name, date, dur = (list(row) + [None] * 3)[:3]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "Treasury Board",
        "notes": "",
    }


def rows_inoria(row, idx):
    name, date, kind, dur = (list(row) + [None] * 4)[:4]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "Inoria",
        "notes": clean_name(kind) if kind and kind != "coaching" else "",
    }


def rows_erickson(row, idx):
    name, date, dur = (list(row) + [None] * 3)[:3]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "Erickson",
        "notes": "",
    }


def rows_pro_bono(row, idx):
    name, date, company, dur = (list(row) + [None] * 4)[:4]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "pro_bono",
        "client_name": clean_name(name),
        "client_organization": clean_name(company),
        "notes": "",
    }


def rows_paid_reciprocoaching(row, idx):
    name, date, dur = (list(row) + [None] * 3)[:3]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "Reciprocoaching",
        "notes": "",
    }


def rows_minimax(row, idx):
    name, date, dur, descr = (list(row) + [None] * 4)[:4]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "Minimax",
        "notes": clean_name(descr),
    }


def rows_eqi_debriefs(row, idx):
    name, company, date, dur, kind = (list(row) + [None] * 5)[:5]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": clean_name(company),
        "assessment_type": clean_name(kind),
        "notes": "",
    }


def rows_kec(row, idx):
    # NAMEs, DATE, Session Duration, descr, Contact
    name, date, dur, descr, contact = (list(row) + [None] * 5)[:5]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "individual",
        "paid_status": "paid",
        "client_name": clean_name(name),
        "client_organization": "KEC",
        "sponsor_contact_name": clean_name(contact),
        "notes": clean_name(descr),
    }


def rows_group_coaching(row, idx):
    # TOPIC, NAMEs, DATE, COMPANY, Session Duration, Contact
    topic, names, date, company, dur, contact = (list(row) + [None] * 6)[:6]
    if not names or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "group",
        "paid_status": "paid",
        "client_name": clean_name(names),
        "client_organization": clean_name(company),
        "sponsor_contact_name": clean_name(contact),
        "notes": clean_name(topic),
    }


def rows_other_group_coaching(row, idx):
    # NAMEs, DATE, Session Duration  (no header row)
    names, date, dur = (list(row) + [None] * 3)[:3]
    if not names or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "session",
        "client_type": "group",
        "paid_status": "paid",
        "client_name": clean_name(names),
        "client_organization": "",
        "notes": "",
    }


def rows_mentor_coaching(row, idx):
    # NAME (mentor receiving from this person), label, DATE, hours, Provider
    name, _label, date, dur, provider = (list(row) + [None] * 5)[:5]
    if not name or not date: return
    h = normalize_hours(dur)
    if h is None: return
    yield {
        "date": normalize_date(date),
        "hours": h,
        "category": "mentor_coaching_received",
        "client_type": "",
        "paid_status": "",
        "client_name": "",
        "client_organization": "",
        "mentor_coach_name": clean_name(name),
        "mentor_coach_organization": clean_name(provider),
        "notes": "",
    }


# ─── Sheet -> handler map ──────────────────────────────────────────────────────

SHEET_HANDLERS = {
    "Paid Private Sessions":    (rows_paid_private,         1),  # skip header rows
    "Treasury Board":           (rows_treasury_board,       2),  # header + blank
    "Inoria":                   (rows_inoria,               0),  # no header
    "Erickson":                 (rows_erickson,             2),
    "Pro Bono Sessions":        (rows_pro_bono,             1),
    "Paid Reciprocoaching":     (rows_paid_reciprocoaching, 2),
    "Minimax":                  (rows_minimax,              2),
    "EQi Debriefs":             (rows_eqi_debriefs,         1),
    "KEC":                      (rows_kec,                  2),
    "Group Coaching":           (rows_group_coaching,       2),
    "Other Group Coaching":     (rows_other_group_coaching, 0),
    "Mentor Coaching":          (rows_mentor_coaching,      0),
    # "TOTAL" — skipped, just a sum.
}


# ─── Main ─────────────────────────────────────────────────────────────────────

OUTPUT_COLUMNS = [
    "date", "hours", "category", "client_type", "paid_status",
    "client_name", "client_organization", "sponsor_contact_name",
    "assessment_type", "mentor_coach_name", "mentor_coach_organization",
    "notes",
]


def main(xlsx_path: Path, csv_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    written = 0
    skipped = 0
    by_sheet: dict[str, int] = {}

    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        w.writeheader()

        for sheet_name in wb.sheetnames:
            if sheet_name not in SHEET_HANDLERS:
                continue
            handler, skip_rows = SHEET_HANDLERS[sheet_name]
            ws = wb[sheet_name]
            count = 0
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx < skip_rows:
                    continue
                # Skip wholly empty rows
                if not any(cell not in (None, "") for cell in row):
                    continue
                emitted = False
                for record in handler(row, idx):
                    if not record["date"] or record["hours"] in (None, 0):
                        skipped += 1
                        continue
                    w.writerow(record)
                    written += 1
                    count += 1
                    emitted = True
                if not emitted:
                    skipped += 1
            by_sheet[sheet_name] = count
            print(f"  {sheet_name:<25s} -> {count:>4d} rows")

    print(f"\nWrote {written} rows, skipped {skipped} blank/invalid")
    print(f"Output: {csv_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    main(Path(sys.argv[1]), Path(sys.argv[2]))
